import os
import time
import sys
import json
import yaml
import boto3
import openai
import requests
import hashlib
import fal_client
from dotenv import load_dotenv
from datetime import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()
fal_key = os.environ.get('FAL_KEY')
openai.api_key = os.environ.get('OPENAI_API_KEY')
finetune_id = os.environ.get('FINE_TUNE_ID')
BUCKET_NAME = "falaiposting"
os.environ["AWS_ACCESS_KEY_ID"] = os.getenv("AWS_ACCESS_KEY_ID")
os.environ["AWS_SECRET_ACCESS_KEY"] = os.getenv("AWS_SECRET_ACCESS_KEY")
os.environ["AWS_REGION"] = os.getenv("AWS_REGION")

required_env = ["FAL_KEY", "OPENAI_API_KEY", "FINE_TUNE_ID", "AWS_ACCESS_KEY_ID", "AWS_SECRET_ACCESS_KEY", "AWS_REGION"]
for var in required_env:
    if not os.getenv(var):
        print(f"ERROR: Required environment variable '{var}' is missing. Please check your .env file.")
        sys.exit(1)


class PromptHandler(FileSystemEventHandler):
    def __init__(self):
        self.prompt_path = os.path.abspath("prompt.yaml")
        self.last_hash = self.get_file_hash()
        self.debounce_seconds = 1.0
        self.last_trigger_time = 0
        self.base_dir = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__))
        self.image_dir = os.path.join(self.base_dir, "images", "approved")
        os.makedirs(self.image_dir, exist_ok=True)
        self.excel_path = os.path.join(self.base_dir, "generated_images.xlsx")
        self.metadata_path = os.path.join(self.image_dir, "metadata.json")

    def get_file_hash(self):
        try:
            with open(self.prompt_path, "rb") as f:
                return hashlib.md5(f.read()).hexdigest()
        except Exception:
            return ""

    def on_modified(self, event):
        if os.path.abspath(event.src_path) != self.prompt_path:
            return
        now = time.time()
        if now - self.last_trigger_time < self.debounce_seconds:
            return
        current_hash = self.get_file_hash()
        if current_hash == self.last_hash:
            return
        self.last_hash = current_hash
        self.last_trigger_time = now
        self.process_prompt()

    def process_prompt(self):
        print("Detected changes to prompt.yaml, processing...\n")
        try:
            with open(self.prompt_path, 'r') as f:
                prompts = yaml.safe_load(f)
            text_prompt = prompts["generate_post"]["user"]
            system_prompt = prompts["generate_post"]["system"]
            image_prompt = prompts["generate_image"]["prompt"]

            print("Calling OpenAI for text...")
            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": text_prompt}
                ]
            )
            generated_text = response.choices[0].message.content.strip()
            print("Text generated\nCalling Fal.ai for images...")

            def on_queue_update(update):
                if isinstance(update, fal_client.InProgress):
                    for log in update.logs:
                        print(log["message"])

            result = fal_client.subscribe(
                "fal-ai/flux-pro/v1.1-ultra-finetuned",
                arguments={
                    "prompt": image_prompt,
                    "guidance_scale": 10,
                    "seed": 0,
                    "sync_mode": False,
                    "num_images": 1,
                    "enable_safety_checker": False,
                    "safety_tolerance": 6,
                    "output_format": "jpeg",
                    "aspect_ratio": "1:1",
                    "finetune_id": finetune_id,
                    "finetune_strength": 0.7
                },
                with_logs=True,
                on_queue_update=on_queue_update
            )

            print("Images generated.\nUpdating Excel...")

            if os.path.exists(self.metadata_path):
                os.remove(self.metadata_path)

            if os.path.exists(self.excel_path):
                wb = load_workbook(self.excel_path)
                if "Image" in wb.sheetnames:
                    wb.remove(wb["Image"])
                ws1 = wb.create_sheet("Image")
                ws2 = wb["Data"] if "Data" in wb.sheetnames else wb.create_sheet("Data")
            else:
                wb = Workbook()
                ws1 = wb.active
                ws1.title = "Image"
                ws2 = wb.create_sheet("Data")
                ws2.append(["URL", "File Name", "Timings", "Prompt", "Generated Text"])

            def style_button(cell, text, color, url):
                cell.value = text
                cell.hyperlink = url
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            image_row = 1
            batch_metadata = {}
            if result and 'images' in result:
                images = [img for img in result['images'] if img['url'].endswith('.jpeg') or img['url'].endswith('.jpg')]
                for idx, image_info in enumerate(images):
                    image_url = image_info['url']
                    file_id = image_url.split('/')[-1].split('.')[0]
                    file_name = f"Aria_{result['seed']}_{file_id}.jpeg"
                    file_path = os.path.join(self.image_dir, file_name)

                    res = requests.get(image_url)
                    if res.status_code != 200:
                        print(f"Failed to download {image_url}")
                        continue
                    with open(file_path, 'wb') as f:
                        f.write(res.content)

                    img = XLImage(file_path)
                    img.width = 793
                    img.height = 497
                    ws1.add_image(img, f"A{image_row}")

                    ws1.cell(row=image_row, column=14).value = generated_text
                    ws1.cell(row=image_row, column=14).alignment = Alignment(wrap_text=True)
                    ws1.column_dimensions['N'].width = 65
                    ws1.row_dimensions[image_row].height = 300

                    approve_url = (
                        f"https://1fs75huc68.execute-api.eu-west-2.amazonaws.com/default/image_authenticator"
                        f"?file={file_name}&url={image_url}&action=approve"
                    )
                    approve_cell = ws1.cell(row=image_row + 1, column=14)
                    style_button(approve_cell, "APPROVE", "00B050", approve_url)

                    row = ws2.max_row + 1
                    ws2.cell(row=row, column=1).value = image_url
                    ws2.cell(row=row, column=2).value = file_name
                    ws2.cell(row=row, column=3, value=str(datetime.now().isoformat()))
                    ws2.cell(row=row, column=4).value = result["prompt"]
                    ws2.cell(row=row, column=5).value = generated_text

                    batch_metadata[file_name] = {
                        "url": image_url,
                        "text": generated_text
                    }

                    image_row += 27

            s3 = boto3.client("s3",aws_access_key_id=os.environ.get("AWS_ACCESS_KEY_ID"),aws_secret_access_key=os.environ.get("AWS_SECRET_ACCESS_KEY"),region_name=os.environ.get("AWS_REGION", "ap-south-1"))
            s3.put_object(
                Bucket=BUCKET_NAME,
                Key="temp/metadata.json",
                Body=json.dumps(batch_metadata),
                ContentType="application/json"
            )
            wb.save(self.excel_path)
            print("Excel updated.\nWaiting for next change...\n")
        except Exception as e:
            print(f"Error: {e}")


if __name__ == "__main__":
    print("Watching for changes in prompt.yaml... Press Ctrl+C to stop.")
    event_handler = PromptHandler()
    observer = Observer()
    observer.schedule(event_handler, path=".", recursive=False)
    observer.start()
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("Stopped watching.")
        observer.stop()
    observer.join()
